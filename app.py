
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

st.title("Efficient Frontier Optimizer")
uploaded_file = st.file_uploader("Upload sector_input.xlsx", type=["xlsx"])

if uploaded_file is not None:
    # Load Excel using openpyxl
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    def read_row(row_num, start_col, end_col):
        return [ws.cell(row=row_num, column=col).value for col in range(start_col, end_col + 1)]

    # Read inputs
    sectors = read_row(3, 2, 17)
    if sectors is None or len(sectors) == 0 or all(s is None for s in sectors):
        st.error("Sectors could not be read. Please ensure row 3, columns B to Q are filled.")
        st.stop()

    expected_returns = np.array(read_row(4, 2, 17), dtype=float)
    volatility = np.array(read_row(5, 2, 17), dtype=float)
    cor_matrix = np.array([read_row(r, 2, 17) for r in range(9, 25)], dtype=float)
    min_weights = np.array(read_row(110, 2, 17), dtype=float)
    max_weights = np.array(read_row(111, 2, 17), dtype=float)
    risk_free_rate = ws["B113"].value or 0.0

    # Validate data dimensions
    n_assets = len(sectors)
    results = np.zeros((4 + n_assets, 10000))

    # Covariance matrix
    D = np.diag(volatility)
    cov_matrix = D @ cor_matrix @ D

    # Monte Carlo simulation
    for i in range(10000):
        weights = np.random.uniform(size=n_assets)
        weights /= np.sum(weights)
        if np.any(weights < min_weights) or np.any(weights > max_weights):
            continue
        port_return = np.dot(weights, expected_returns)
        port_stddev = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
        sharpe_ratio = (port_return - risk_free_rate) / port_stddev if port_stddev != 0 else 0
        results[0, i] = port_return
        results[1, i] = port_stddev
        results[2, i] = sharpe_ratio
        results[3:, i] = weights

    results_df = pd.DataFrame(results.T, columns=["Return", "Risk", "Sharpe"] + sectors)
    results_df = results_df[(results_df["Return"] > 0) & (results_df["Risk"] > 0)]

    max_sharpe_port = results_df.iloc[results_df["Sharpe"].idxmax()]
    min_risk_port = results_df.iloc[results_df["Risk"].idxmin()]

    st.subheader("Max Sharpe Portfolio Allocation")
    st.dataframe(max_sharpe_port[3:].apply(lambda x: f"{x:.2%}"))

    st.subheader("Min Risk Portfolio Allocation")
    st.dataframe(min_risk_port[3:].apply(lambda x: f"{x:.2%}"))

    fig, ax = plt.subplots()
    scatter = ax.scatter(results_df["Risk"], results_df["Return"], c=results_df["Sharpe"], cmap="viridis", alpha=0.7)
    ax.scatter(max_sharpe_port["Risk"], max_sharpe_port["Return"], c="red", marker="*", s=200, label="Max Sharpe")
    ax.scatter(min_risk_port["Risk"], min_risk_port["Return"], c="blue", marker="X", s=100, label="Min Risk")
    ax.set_xlabel("Volatility (Risk)")
    ax.set_ylabel("Expected Return")
    ax.set_title("Efficient Frontier with Monte Carlo Simulation")
    ax.legend()
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.2%}"))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.2%}"))
    st.pyplot(fig)
